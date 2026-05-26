"""Phase 3 — Summary Excel generator.

Produces KSyder_Summary_<date_from>_to_<date_to>_<ts>.xlsx with two sheets:
  • Summary  — one row per site: counts, category breakdown, crawl status
  • Meta     — run-level totals and parameters
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger("gov_aggregator.exporters.excel_summary")

# ── openpyxl ──────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:  # pragma: no cover
    _OPENPYXL_OK = False

# ── Palette ───────────────────────────────────────────────────────────────────
_NAVY    = "1e3a5f"
_GRAY_H  = "d9d9d9"   # ministry-group header row
_GRAY_F  = "bfbfbf"   # footer totals row
_RED_BG  = "ffcccc"   # 0 items in range
_YEL_BG  = "fff3cc"   # 1-5 items
_GRN_BG  = "ccffcc"   # >5 items
_FAIL_FG = "c00000"   # failed status text
_SUCC_FG = "1a6e27"   # success/completed text
_CACH_FG = "0563c1"   # cached text

_CATEGORIES = ["circular", "tender", "recruitment", "notification", "news"]

_HEADERS = [
    "Ministry",
    "Site Name",
    "Site Key",
    "Total Items Found",
    "Items in Date Range",
    *[c.capitalize() for c in _CATEGORIES],
    "Has PDF",
    "Crawl Status",
    "Error Message",
    "Crawl Time",
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> "Border":
    s = Side(style="thin", color="cccccc")
    return Border(left=s, right=s, top=s, bottom=s)


def _passes_range(item: dict[str, Any], date_from: str | None, date_to: str | None) -> bool:
    """True when item falls inside the requested date range (or range is open-ended)."""
    pd = item.get("publish_date")
    if not pd:
        return True
    try:
        dt = datetime.fromisoformat(pd)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if date_from:
            df = datetime.fromisoformat(date_from + "T00:00:00+00:00")
            if dt < df:
                return False
        if date_to:
            de = datetime.fromisoformat(date_to + "T23:59:59+00:00")
            if dt > de:
                return False
        return True
    except (ValueError, TypeError):
        return True


def _auto_fit(ws: Any, min_w: int = 12, max_w: int = 50) -> None:
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value or "")) for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 2, max_w))


# ── Core builder ─────────────────────────────────────────────────────────────

def _build_site_rows(
    crawl_result: dict[str, Any],
    date_from: str | None,
    date_to: str | None,
) -> list[dict[str, Any]]:
    """Aggregate items and statuses into one dict per site."""
    items: list[dict[str, Any]] = crawl_result.get("items", [])
    statuses: list[dict[str, Any]] = crawl_result.get("site_statuses", [])

    # Build per-site item buckets from the already-date-filtered items list
    items_by_site: dict[str, list[dict[str, Any]]] = {}
    for item in items:
        sk = item.get("site_key", "")
        items_by_site.setdefault(sk, []).append(item)

    status_by_site: dict[str, dict[str, Any]] = {s["site_key"]: s for s in statuses}

    # Union of all known site keys
    all_keys: list[str] = list(
        dict.fromkeys(
            [s["site_key"] for s in statuses] + list(items_by_site.keys())
        )
    )

    rows: list[dict[str, Any]] = []
    for sk in all_keys:
        st = status_by_site.get(sk, {})
        site_items = items_by_site.get(sk, [])

        # items arriving here already passed the server-side date filter;
        # count all of them as "in range" as well (they already are)
        in_range = [i for i in site_items if _passes_range(i, date_from, date_to)]

        cat_counts: dict[str, int] = {c: 0 for c in _CATEGORIES}
        for item in in_range:
            cat = item.get("category", "news").lower()
            if cat in cat_counts:
                cat_counts[cat] += 1

        has_pdf = any(item.get("is_pdf") for item in in_range)

        state = st.get("state", "unknown")
        crawl_time_raw = st.get("crawl_time") or crawl_result.get("crawl_time", "")
        try:
            crawl_time = datetime.fromisoformat(crawl_time_raw).strftime("%Y-%m-%d %H:%M:%S") if crawl_time_raw else ""
        except ValueError:
            crawl_time = crawl_time_raw

        rows.append(
            {
                "ministry":    st.get("ministry") or "",
                "site_name":   st.get("site_name") or sk,
                "site_key":    sk,
                "total":       len(site_items),
                "in_range":    len(in_range),
                "cat_counts":  cat_counts,
                "has_pdf":     "Yes" if has_pdf else "No",
                "state":       state,
                "error":       st.get("message", "") if state in {"error", "missing"} else "",
                "crawl_time":  crawl_time,
            }
        )

    # Sort: ministry → site_name
    rows.sort(key=lambda r: (r["ministry"].lower(), r["site_name"].lower()))
    return rows


def _write_summary_sheet(
    ws: Any,
    rows: list[dict[str, Any]],
) -> None:
    # ── Header row ────────────────────────────────────────────────────────────
    ws.append(_HEADERS)
    hdr_fill = _fill(_NAVY)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    border = _thin_border()

    # ── Data rows ─────────────────────────────────────────────────────────────
    totals: dict[str, int] = {
        "total": 0, "in_range": 0,
        **{c: 0 for c in _CATEGORIES},
    }

    current_ministry = None
    row_num = 2
    for row_data in rows:
        ministry = row_data["ministry"] or "—"
        if ministry != current_ministry:
            current_ministry = ministry

        row_vals = [
            ministry,
            row_data["site_name"],
            row_data["site_key"],
            row_data["total"],
            row_data["in_range"],
            *[row_data["cat_counts"][c] for c in _CATEGORIES],
            row_data["has_pdf"],
            row_data["state"],
            row_data["error"],
            row_data["crawl_time"],
        ]
        ws.append(row_vals)

        # Accumulate totals
        totals["total"]    += row_data["total"]
        totals["in_range"] += row_data["in_range"]
        for c in _CATEGORIES:
            totals[c] += row_data["cat_counts"][c]

        # Apply row formatting
        data_row = ws[row_num]
        in_range_val = row_data["in_range"]
        if in_range_val == 0:
            range_fill = _fill(_RED_BG)
        elif in_range_val <= 5:
            range_fill = _fill(_YEL_BG)
        else:
            range_fill = _fill(_GRN_BG)

        state = row_data["state"]
        if state in {"error", "missing", "failed"}:
            status_color = _FAIL_FG
        elif state in {"completed", "success"}:
            status_color = _SUCC_FG
        elif state == "cached":
            status_color = _CACH_FG
        else:
            status_color = "000000"

        for i, cell in enumerate(data_row):
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            # Col E (index 4) = "Items in Date Range" — colored
            if i == 4:
                cell.fill = range_fill
                cell.font = Font(bold=True)
            # Col L (index 11) = "Crawl Status" — colored text
            if i == 11:
                cell.font = Font(color=status_color, bold=(state in {"error", "missing"}))

        row_num += 1

    # ── Footer totals row ─────────────────────────────────────────────────────
    footer = [
        "TOTAL", "", "",
        totals["total"],
        totals["in_range"],
        *[totals[c] for c in _CATEGORIES],
        "", "", "", "",
    ]
    ws.append(footer)
    footer_row = ws[row_num]
    footer_fill = _fill(_GRAY_F)
    footer_font = Font(bold=True, size=11)
    for cell in footer_row:
        cell.fill = footer_fill
        cell.font = footer_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _auto_fit(ws)
    ws.sheet_properties.tabColor = _NAVY


def _write_meta_sheet(
    ws: Any,
    crawl_result: dict[str, Any],
    rows: list[dict[str, Any]],
    date_from: str | None,
    date_to: str | None,
    generated_at: str,
) -> None:
    ws.sheet_properties.tabColor = "808080"

    meta = crawl_result.get("meta", {})
    total_sites = meta.get("requested_sites", len(rows))
    sites_with_updates = sum(1 for r in rows if r["in_range"] > 0)
    sites_no_updates   = sum(1 for r in rows if r["in_range"] == 0 and r["state"] not in {"error", "missing", "unsupported"})
    sites_failed       = sum(1 for r in rows if r["state"] in {"error", "missing"})
    total_items        = sum(r["total"]    for r in rows)
    total_in_range     = sum(r["in_range"] for r in rows)

    date_range_label = f"{date_from or 'N/A'} to {date_to or 'N/A'}"

    fields = [
        ("Generated At",          generated_at),
        ("Date Range",            date_range_label),
        ("Total Sites Crawled",   total_sites),
        ("Sites With Updates",    sites_with_updates),
        ("Sites With No Updates", sites_no_updates),
        ("Sites Failed",          sites_failed),
        ("Total Items Found",     f"{total_items:,}"),
        ("Total Items in Range",  f"{total_in_range:,}"),
    ]

    hdr_fill = _fill(_NAVY)
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    val_font = Font(size=11)
    border = _thin_border()

    for field_name, value in fields:
        ws.append([field_name, value])
        r = ws[ws.max_row]
        r[0].fill = hdr_fill
        r[0].font = hdr_font
        r[1].font = val_font
        r[0].border = border
        r[1].border = border
        r[0].alignment = Alignment(horizontal="left", vertical="center")
        r[1].alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36


# ── Public API ────────────────────────────────────────────────────────────────

async def generate_summary_excel(
    crawl_result: dict[str, Any],
    date_from: str | None,
    date_to: str | None,
    output_path: str,
) -> str:
    """Generate a summary Excel workbook and save it to *output_path*.

    Returns the resolved absolute path as a string.
    """
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl>=3.1.0")

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    rows = _build_site_rows(crawl_result, date_from, date_to)

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_meta = wb.create_sheet("Meta")

    _write_summary_sheet(ws_summary, rows)
    _write_meta_sheet(ws_meta, crawl_result, rows, date_from, date_to, generated_at)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Summary Excel saved: %s (%d sites)", out, len(rows))
    return str(out.resolve())
