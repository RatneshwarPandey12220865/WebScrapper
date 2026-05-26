"""Phase 4 — Per-site detailed Excel generator.

Produces one .xlsx per site with a 5-row header block followed by a
formatted item table.  Columns: #, Title, Category, Published Date,
Crawl Time, Section, Description, Document Link, Source Page, Is PDF.
"""
from __future__ import annotations

import asyncio
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

logger = logging.getLogger("gov_aggregator.exporters.excel_site_detail")

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:  # pragma: no cover
    _OPENPYXL_OK = False

# ── Palette ───────────────────────────────────────────────────────────────────
_NAVY      = "1e3a5f"
_HDR_BLUE  = "d6e4f7"   # column-header row background
_ALT_ROW   = "f8f8f8"   # alternating row tint

_CAT_FILL: dict[str, str] = {
    "circular":     "ddeeff",
    "tender":       "ffe8cc",
    "recruitment":  "d4edda",
    "notification": "fff3cc",
    "news":         "eeeeee",
}

_COL_HEADERS = [
    "#", "Title", "Category", "Published Date",
    "Crawl Time", "Section", "Description",
    "Document Link", "Source Page", "Is PDF",
]

# Column widths (characters)
_COL_WIDTHS = [5, 50, 14, 16, 18, 18, 60, 20, 36, 8]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> "Border":
    s = Side(style="thin", color="cccccc")
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_date(iso_str: str | None) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d-%b-%Y")
    except (ValueError, TypeError):
        return iso_str


def _fmt_datetime(iso_str: str | None) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d-%b-%Y %H:%M")
    except (ValueError, TypeError):
        return iso_str


def _slugify_ministry(ministry: str) -> str:
    """Convert ministry name to a safe filename prefix (max 30 chars)."""
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", ministry).strip("_")
    return slug[:30] if slug else "Unknown"


def _passes_range(item: dict[str, Any], date_from: str | None, date_to: str | None) -> bool:
    pd = item.get("publish_date")
    if not pd:
        return True
    try:
        dt = datetime.fromisoformat(pd)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        if date_from:
            df = datetime.fromisoformat(date_from + "T00:00:00+00:00")
            if dt < df:
                return False
        if date_to:
            de = datetime.fromisoformat(date_to + "T23:59:59+00:00")
            if dt > de:
                return False
        return True
    except (ValueError, TypeError):
        return True


# ── Sheet builder ─────────────────────────────────────────────────────────────

def _write_site_sheet(
    ws: Any,
    items: list[dict[str, Any]],
    *,
    ministry: str,
    site_name: str,
    site_key: str,
    date_from: str | None,
    date_to: str | None,
    crawl_time: str | None,
) -> None:
    border = _thin_border()
    navy_fill = _fill(_NAVY)
    navy_font = Font(bold=True, color="FFFFFF", size=11)

    # ── Header block rows 1-5 ────────────────────────────────────────────────
    last_col = len(_COL_HEADERS)
    last_col_letter = get_column_letter(last_col)

    def _header_row(label: str, value: str, row: int) -> None:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        # Merge value across remaining columns
        ws.merge_cells(f"B{row}:{last_col_letter}{row}")
        lbl = ws.cell(row, 1)
        lbl.fill = navy_fill
        lbl.font = navy_font
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        lbl.border = border
        val = ws.cell(row, 2)
        val.font = Font(bold=True, size=11)
        val.alignment = Alignment(horizontal="left", vertical="center")
        val.border = border
        ws.row_dimensions[row].height = 20

    date_range_label = f"{date_from or 'N/A'}  →  {date_to or 'N/A'}"
    crawl_label = _fmt_datetime(crawl_time) if crawl_time else "N/A"

    _header_row("Ministry:",   ministry,         1)
    _header_row("Site:",       site_name,        2)
    _header_row("Date Range:", date_range_label, 3)
    _header_row("Items:",      f"{len(items)}  |  Crawl Time: {crawl_label}", 4)

    # Row 5: blank separator
    ws.row_dimensions[5].height = 6

    # ── Column header row 6 ───────────────────────────────────────────────────
    hdr_fill = _fill(_HDR_BLUE)
    hdr_font = Font(bold=True, size=10)
    ws.append([])          # skip row 5 (blank)
    ws.append(_COL_HEADERS)
    hdr_row = ws[6]
    for cell in hdr_row:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[6].height = 18

    ws.freeze_panes = "A7"

    # ── Data rows starting at row 7 ───────────────────────────────────────────
    for i, item in enumerate(items, start=1):
        cat   = (item.get("category") or "news").lower()
        link  = item.get("link") or ""
        pdf   = item.get("pdf_url") or link if item.get("is_pdf") else ""
        ext   = item.get("external_link") or link if not item.get("is_pdf") else ""
        doc_url = pdf or ext or link

        row_vals = [
            i,
            item.get("title") or "",
            cat.capitalize(),
            _fmt_date(item.get("publish_date")),
            _fmt_datetime(item.get("crawl_time")),
            item.get("section_label") or "",
            item.get("description") or "",
            "",          # Document Link — set as hyperlink below
            item.get("crawl_url") or "",
            "Yes" if item.get("is_pdf") else "No",
        ]
        ws.append(row_vals)
        data_row_num = 6 + i
        data_row = ws[data_row_num]

        # Alternating row background
        row_fill = _fill(_ALT_ROW) if i % 2 == 0 else None
        cat_fill = _fill(_CAT_FILL.get(cat, "ffffff"))

        for j, cell in enumerate(data_row):
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if row_fill:
                cell.fill = row_fill
            # Category cell (col C = index 2) — colour by category
            if j == 2:
                cell.fill = cat_fill
                cell.font = Font(bold=True, size=10)
            # Title (col B = index 1) — wrap text
            if j == 1:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Description (col G = index 6) — wrap, max height 60
            if j == 6:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            # Is PDF (col J = index 9) — centered
            if j == 9:
                cell.alignment = Alignment(horizontal="center", vertical="top")

        # Document Link hyperlink (col H = column 8)
        link_cell = ws.cell(data_row_num, 8)
        if doc_url:
            link_cell.hyperlink = doc_url
            link_cell.value = "Open Document"
            link_cell.font = Font(color="0563c1", underline="single", size=10)
        else:
            link_cell.value = ""

        ws.row_dimensions[data_row_num].height = 40

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Col A header block — narrow label
    ws.column_dimensions["A"].width = 14

    # ── Print settings ────────────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_area = f"A1:{last_col_letter}{6 + len(items)}"
    ws.sheet_properties.tabColor = _NAVY


# ── Public API ────────────────────────────────────────────────────────────────

def _safe_filename(ministry: str, site_key: str, date_from: str | None) -> str:
    min_slug = _slugify_ministry(ministry)
    date_slug = (date_from or "open").replace("-", "")
    return f"{min_slug}_{site_key}_{date_slug}.xlsx"


async def generate_site_detail_excel(
    site_key: str,
    items: list[dict[str, Any]],
    status: dict[str, Any],
    date_from: str | None,
    date_to: str | None,
    output_path: str,
) -> str:
    """Generate a per-site detail workbook and save it to *output_path*.

    Returns the resolved absolute path.
    """
    if not _OPENPYXL_OK:
        raise RuntimeError("openpyxl is not installed.")

    ministry   = status.get("ministry") or "Unknown Ministry"
    site_name  = status.get("site_name") or site_key
    crawl_time = status.get("crawl_time")

    # Filter to items in range (items already passed server-side filter,
    # but apply UI range as a second pass when date_from/date_to are set)
    in_range = [i for i in items if _passes_range(i, date_from, date_to)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    _write_site_sheet(
        ws,
        in_range,
        ministry=ministry,
        site_name=site_name,
        site_key=site_key,
        date_from=date_from,
        date_to=date_to,
        crawl_time=crawl_time,
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    await asyncio.to_thread(wb.save, out)
    logger.info("Site detail Excel saved: %s (%d items)", out, len(in_range))
    return str(out.resolve())
