"""Generate KSyder Site Status Excel report."""
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

BASE = Path(__file__).parent / "gov_aggregator" / "data"

# Load known_sites.json
with open(BASE / "known_sites.json", encoding="utf-8") as f:
    known = json.load(f)["sites"]

# Custom crawler site keys (from scrapers/custom/__init__.py)
CUSTOM_KEYS = {
    "archaeological-survey-of-india",
    "cbic-customs",
    "cci",
    "competition-commission-of-india",
    "chemexcil",
    "chandigarh",
    "industry-and-internal-trade",
    "department-of-land-resources",
    "epfo",
    "minisry-of-commerce",
    "department-of-bio-technology",
    "department-of-health-research",
    "directorate-general-of-foreign-trade",
    "dot",
    "fssai",
    "gst",
    "income-tax",
    "irdai",
    "ministry-of-labour",
    "ministry-of-road-transport-and-highways",
    "meity",
    "ministry-of-ayush",
    "national-centre-for-cold-chain-development",
    "national-medical-commission",
    "nse",
    "pharmacy-council",
    "department-of-pharmaceuticals",
    "power-ministry",
    "power-ministry-pib",
    "rajasthan",
    "rbi",
    "coffee-board",
    "department-of-fertilizers",
    "department-of-fisheries",
    "icmr",
    "sebi",
    "employees-provident-fund-organisation",
    "indian-council-of-medical-research",
    "national-medical-commission",
    "pharmacy-council",
}

def scraping_type(site):
    key = site.get("site_key", "")
    status = site.get("status", "unknown")
    status_raw = site.get("status_raw") or ""

    if status == "not_working":
        return "Not Working / Blocked"
    if status == "unknown" or (
        site.get("preferred_url") in (None, "N/A", "working", "") and
        site.get("registry_url") in ("", "N/A", "working", None)
    ):
        return "Unknown / Not Configured"
    if key in CUSTOM_KEYS:
        return "Custom Scraper"
    # If status is working
    return "Normal Scraping"

def best_url(site):
    pref = site.get("preferred_url") or ""
    reg  = site.get("registry_url") or ""
    alt  = site.get("alternate_url") or ""
    for u in [pref, reg, alt]:
        if u and u not in ("N/A", "working", "not working", ""):
            return u
    return ""

# Build rows
rows = []
for s in known:
    name = s.get("name", "")
    url  = best_url(s)
    stype = scraping_type(s)
    rows.append((name, url, stype))

# Sort exactly like the KSyder UI:
#   key = (not supported, name.lower())
#   supported sites (Custom or Normal) first alphabetically,
#   then unsupported / unknown alphabetically.
UNSUPPORTED_TYPES = {"Not Working / Blocked", "Unknown / Not Configured"}
rows.sort(key=lambda r: (r[2] in UNSUPPORTED_TYPES, r[0].lower()))

# ---- Excel styling helpers ----
def thin_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

TYPE_FILLS = {
    "Custom Scraper":           PatternFill("solid", fgColor="FFF2CC"),   # light yellow
    "Normal Scraping":          PatternFill("solid", fgColor="E2EFDA"),   # light green
    "Not Working / Blocked":    PatternFill("solid", fgColor="FCE4D6"),   # light red
    "Unknown / Not Configured": PatternFill("solid", fgColor="EDEDED"),   # light grey
}
STATUS_FILLS = {
    "Not Done": PatternFill("solid", fgColor="FF6B6B"),   # red
    "Done":     PatternFill("solid", fgColor="70AD47"),   # green
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Site Status"

# ---- Header ----
headers = ["S.No", "Site Name", "Site Link", "Scraping Type", "Status", "Description"]
header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

ws.row_dimensions[1].height = 30

# ---- Data rows ----
for i, (name, url, stype) in enumerate(rows, 1):
    row_num = i + 1
    # S.No
    c = ws.cell(row=row_num, column=1, value=i)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

    # Site Name
    c = ws.cell(row=row_num, column=2, value=name)
    c.fill = TYPE_FILLS.get(stype, PatternFill())
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = thin_border()

    # Site Link (hyperlink)
    c = ws.cell(row=row_num, column=3, value=url)
    if url and url.startswith("http"):
        c.hyperlink = url
        c.font = Font(color="0563C1", underline="single")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = thin_border()

    # Scraping Type
    c = ws.cell(row=row_num, column=4, value=stype)
    c.fill = TYPE_FILLS.get(stype, PatternFill())
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.font = Font(bold=True, size=10)
    c.border = thin_border()

    # Status — "Not Done" highlighted RED
    c = ws.cell(row=row_num, column=5, value="Not Done")
    c.fill = STATUS_FILLS["Not Done"]
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border()

    # Description (blank for now)
    c = ws.cell(row=row_num, column=6, value="")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.border = thin_border()

    ws.row_dimensions[row_num].height = 22

# ---- Column widths ----
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 58
ws.column_dimensions["D"].width = 26
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 50

# ---- Freeze panes (header + S.No) ----
ws.freeze_panes = "B2"

# ---- Auto-filter ----
ws.auto_filter.ref = f"A1:F{len(rows)+1}"

# ---- Legend sheet ----
ls = wb.create_sheet("Legend")
ls.column_dimensions["A"].width = 28
ls.column_dimensions["B"].width = 50
legend_data = [
    ("Color / Value", "Meaning"),
    ("Yellow row",     "Custom Scraper — uses a dedicated async crawler (Playwright / API)"),
    ("Green row",      "Normal Scraping — standard HTTP + BeautifulSoup / Scrapy selectors"),
    ("Red row",        "Not Working / Blocked — anti-bot, broken, or unreachable"),
    ("Grey row",       "Unknown / Not Configured — no URL or scraper yet"),
    ("Status = Not Done (red)",  "Site description not yet reviewed"),
    ("Status = Done (green)",    "Site description filled in and reviewed"),
]
for r, (a, b) in enumerate(legend_data, 1):
    ls.cell(row=r, column=1, value=a).font = Font(bold=(r==1))
    ls.cell(row=r, column=2, value=b)

out = Path(__file__).parent / "KSyder_Site_Status.xlsx"
wb.save(out)
print(f"Saved: {out}  ({len(rows)} sites)")
